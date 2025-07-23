import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io
import tempfile
import os
from datetime import datetime

class ExactPackagingTemplateManager:
    def __init__(self):
        self.template_fields = {
            # Header Information
            'Revision No.': '',
            'Date': '',
            
            # Vendor Information
            'Vendor Code': '',
            'Vendor Name': '',
            'Vendor Location': '',
            
            # Part Information
            'Part No.': '',
            'Part Description': '',
            'Part Unit Weight': '',
            'Part Weight Unit': '',
            'Part L': '',
            'Part W': '',
            'Part H': '',
            
            # Inner Packaging (distinct from Primary)
            'Inner Packaging Type': '',
            'Inner L': '',
            'Inner W': '',
            'Inner H': '',
            'Inner Qty/Pack': '',
            'Inner Empty Weight': '',
            'Inner Pack Weight': '',
            
            # Primary Packaging
            'Primary Packaging Type': '',
            'Primary L-mm': '',
            'Primary W-mm': '',
            'Primary H-mm': '',
            'Primary Qty/Pack': '',
            'Primary Empty Weight': '',
            'Primary Pack Weight': '',
            
            # Secondary Packaging
            'Secondary Packaging Type': '',
            'Secondary L-mm': '',
            'Secondary W-mm': '',
            'Secondary H-mm': '',
            'Secondary Qty/Pack': '',
            'Secondary Empty Weight': '',
            'Secondary Pack Weight': '',
            
            # Packaging Procedures (10 steps)
            'Procedure Step 1': '',
            'Procedure Step 2': '',
            'Procedure Step 3': '',
            'Procedure Step 4': '',
            'Procedure Step 5': '',
            'Procedure Step 6': '',
            'Procedure Step 7': '',
            'Procedure Step 8': '',
            'Procedure Step 9': '',
            'Procedure Step 10': '',
            'Procedure Step 11': '',
            
            # Approval
            'Issued By': '',
            'Reviewed By': '',
            'Approved By': '',
            
            # Additional fields
            'Problem If Any': '',
            'Caution': ''
        }
        
        # Predefined packaging procedures for different types
        self.packaging_procedures = {
            "BOX IN BOX SENSITIVE": [
                "Pick up 1 quantity of part and apply bubble wrapping over it",
                "Apply tape and Put 1 such bubble wrapped part into a carton box [L-{Inner L} mm, W-{Inner W} mm, H-{Inner H} mm]",
                "Seal carton box and put {Inner Qty/Pack} such carton boxes into another carton box [L-{Inner L} mm, W-{Inner W} mm, H-{Inner H} mm]",
                "Seal carton box and put Traceability label as per PMSPL standard guideline",
                "Prepare additional carton boxes in line with procurement schedule (multiple of pack quantity -- {Inner Qty/Pack})",
                "If procurement schedule is for less no. of boxes, then load similar boxes of other parts on same wooden pallet.",
                "Load carton boxes on base wooden pallet -- {Layer} boxes per layer & max {Level} level (max height including pallet -1000 mm)",
                "Put corner / edge protector and apply pet strap (2 times -- cross way)",
                "Apply traceability label on complete pack",
                "Attach packing list along with dispatch document and tag copy of same on pack (in case of multiple parts on same pallet)",
                "Ensure Loading/Unloading of palletize load using Hand pallet / stacker / forklift only."
            ],
            
            "BOX IN BOX": [
                "Pick up 1 quantity of part and put it in a polybag",
                "seal the polybag and put it into a carton box [L-{Inner L} mm, W-{Inner W} mm, H-{Inner H} mm]",
                "Put {Inner Qty/Pack} such carton boxes into another carton box [L-{Inner L} mm, W-{Inner W} mm, H-{Inner H} mm]",
                "Seal carton box and put Traceability label as per PMSPL standard guideline",
                "Prepare additional carton boxes in line with procurement schedule (multiple of pack quantity -- {Inner Qty/Pack})",
                "If procurement schedule is for less no. of boxes, then load similar boxes of other parts on same wooden pallet.",
                "Load carton boxes on base wooden pallet -- {Layer} boxes per layer & max {Level} level (max height including pallet -1000 mm)",
                "Put corner / edge protector and apply pet strap (2 times -- cross way)",
                "Apply traceability label on complete pack",
                "Attach packing list along with dispatch document and tag copy of same on pack (in case of multiple parts on same pallet)",
                "Ensure Loading/Unloading of palletize load using Hand pallet / stacker / forklift only."
            ],
            
            "CARTON BOX WITH SEPARATOR FOR ONE PART": [
                "Pick up {Qty/Veh} parts and apply bubble wrapping over it (individually)",
                "Apply tape and Put bubble wrapped part into a carton box. Apply part separator & filler material between two parts to arrest part movement during handling",
                "Seal carton box and put Traceability label as per PMSPL standard guideline",
                "Prepare additional carton boxes in line with procurement schedule (multiple of pack quantity -- {Inner Qty/Pack})",
                "Load carton boxes on base wooden pallet -- {Layer} boxes per layer & max {Level} level",
                "If procurement schedule is for less no. of boxes, then load similar boxes of other parts on same wooden pallet.",
                "Put corner / edge protector and apply pet strap (2 times -- cross way)",
                "Apply traceability label on complete pack",
                "Attach packing list along with dispatch document and tag copy of same on pack (in case of multiple parts on same pallet)",
                "Ensure Loading/Unloading of palletize load using Hand pallet / stacker / forklift only.",
                ""
            ],
            
            "INDIVIDUAL NOT SENSITIVE": [
                "Pick up one part and put it into a polybag",
                "Seal polybag and Put polybag into a carton box",
                "Seal carton box and put Traceability label as per PMSPL standard guideline",
                "Prepare additional carton boxes in line with procurement schedule (multiple of pack quantity -- {Inner Qty/Pack})",
                "Load carton boxes on base wooden pallet -- Maximum {Layer} boxes per layer & Maximum {Level} level (max height including pallet - 1000 mm)",
                "If procurement schedule is for less no. of boxes, then load similar boxes of other parts on same wooden pallet.",
                "Put corner / edge protector and apply pet strap (2 times -- cross way)",
                "Apply traceability label on complete pack",
                "Attach packing list along with dispatch document and tag copy of same on pack (in case of multiple parts on same pallet)",
                "Ensure Loading/Unloading of palletize load using Hand pallet / stacker / forklift only.",
                ""
            ],
            
            "INDIVIDUAL PROTECTION FOR EACH PART": [
                "Pick up {Qty/Veh} parts and apply bubble wrapping over it (individually)",
                "Apply tape and Put bubble wrapped part into a carton box. Apply part separator & filler material between two parts to arrest part movement during handling",
                "Seal carton box and put Traceability label as per PMSPL standard guideline",
                "Prepare additional carton boxes in line with procurement schedule (multiple of pack quantity -- {Inner Qty/Pack})",
                "Load carton boxes on base wooden pallet -- {Layer} boxes per layer & max {Level} level (max height including pallet - 1000 mm)",
                "If procurement schedule is for less no. of boxes, then load similar boxes of other parts on same wooden pallet.",
                "Put corner / edge protector and apply pet strap (2 times -- cross way)",
                "Apply traceability label on complete pack",
                "Attach packing list along with dispatch document and tag copy of same on pack (in case of multiple parts on same pallet)",
                "Ensure Loading/Unloading of palletize load using Hand pallet / stacker / forklift only.",
                ""
            ],
            
            "INDIVIDUAL SENSITIVE": [
                "Pick up one part and apply bubble wrapping over it",
                "Apply tape and Put bubble wrapped part into a carton box",
                "Seal carton box and put Traceability label as per PMSPL standard guideline",
                "Prepare additional carton boxes in line with procurement schedule (multiple of pack quantity -- {Inner Qty/Pack})",
                "Load carton boxes on base wooden pallet -- {Layer} boxes per layer & max {Level} level (max height including pallet - 1000 mm)",
                "If procurement schedule is for less no. of boxes, then load similar boxes of other parts on same wooden pallet.",
                "Put corner / edge protector and apply pet strap (2 times -- cross way)",
                "Apply traceability label on complete pack",
                "Attach packing list along with dispatch document and tag copy of same on pack (in case of multiple parts on same pallet)",
                "Ensure Loading/Unloading of palletize load using Hand pallet / stacker / forklift only.",
                ""
            ],
            
            "MANY IN ONE TYPE": [
                "Pick up {Qty/Veh} quantity of part and put it in a polybag",
                "Seal polybag and Put it into a carton box",
                "Seal carton box and put Traceability label as per PMSPL standard guideline",
                "Prepare additional carton boxes in line with procurement schedule (multiple of pack quantity -- {Inner Qty/Pack})",
                "If procurement schedule is for less no. of boxes, then load similar boxes of other parts on same wooden pallet.",
                "Load carton boxes on base wooden pallet -- {Layer} boxes per layer & max {Level} level (max height including pallet - 1000 mm)",
                "Put corner / edge protector and apply pet strap (2 times -- cross way)",
                "Apply traceability label on complete pack",
                "Attach packing list along with dispatch document and tag copy of same on pack (in case of multiple parts on same pallet)",
                "Ensure Loading/Unloading of palletize load using Hand pallet / stacker / forklift only.",
                ""
            ],
            
            "SINGLE BOX": [
                "Pick up 1 quantity of part and put it in a polybag",
                "Put into a carton box",
                "Seal carton box and put Traceability label as per PMSPL standard guideline",
                "Prepare additional carton boxes in line with procurement schedule (multiple of pack quantity -- {Inner Qty/Pack})",
                "If procurement schedule is for less no. of boxes, then load similar boxes of other parts on same wooden pallet.",
                "Load carton boxes on base wooden pallet -- {Layer} boxes per layer & max {Level} level",
                "Put corner / edge protector and apply pet strap (2 times -- cross way) and stretch wrap it",
                "Apply traceability label on complete pack",
                "Attach packing list along with dispatch document and tag copy of same on pack (in case of multiple parts on same pallet)",
                "Ensure Loading/Unloading of palletize load using Hand pallet / stacker / forklift only.",
                ""
            ]
        }
    
    def get_procedure_steps(self, packaging_type, data_dict=None):
        """Get predefined procedure steps for selected packaging type with placeholders filled"""
        procedures = self.packaging_procedures.get(packaging_type, [""] * 11)
        if data_dict:
            # Fill in placeholders with actual values
            filled_procedures = []
            for procedure in procedures:
                filled_procedure = procedure
            
                # Define replacements - ONLY Inner dimensions and basic parameters
                replacements = {
                    # ONLY Inner dimensions (no Primary/Secondary references in procedures)
                    # ONLY Inner dimensions (no Primary/Secondary references in procedures)
                    '{Inner L}': str(data_dict.get('Inner L', 'XXX')),
                    '{Inner W}': str(data_dict.get('Inner W', 'XXX')),  
                    '{Inner H}': str(data_dict.get('Inner H', 'XXX')),
                    '{Inner Qty/Pack}': str(data_dict.get('Inner Qty/Pack', 'XXX')),
                    
                    # Generic quantities (for backwards compatibility)
                    '{Qty/Pack}': str(data_dict.get('Inner Qty/Pack', data_dict.get('Qty/Pack', 'XXX'))),
                    
                    # Other parameters
                    '{Qty/Veh}': str(data_dict.get('Qty/Veh', 'XXX')),
                    '{Layer}': str(data_dict.get('Layer', 'XXX')),
                    '{Level}': str(data_dict.get('Level', 'XXX')),
                }
            
                # Apply all replacements
                for placeholder, value in replacements.items():
                    if placeholder in filled_procedure:
                        filled_procedure = filled_procedure.replace(placeholder, value)
                        print(f"Replaced {placeholder} with {value}")
                        
                filled_procedures.append(filled_procedure)
                
            # Debug: Print what placeholders were found and replaced
            print("\n=== PLACEHOLDER REPLACEMENT DEBUG ===")
            for i, (original, filled) in enumerate(zip(procedures, filled_procedures)):
                if original != filled:
                    print(f"Step {i+1} had replacements:")
                    print(f"  Original: {original}")
                    print(f"  Filled: {filled}")
            print("=====================================\n")
            return filled_procedures
        else:
            return procedures
            
    def extract_data_from_excel(self, uploaded_file):
        """Extract data from uploaded Excel file"""
        extracted_data = {}
        try:
            # Read the Excel file
            df = pd.read_excel(uploaded_file, sheet_name=0)
            
            # Create a mapping of possible column names to our field names
            field_mapping = {
                # Basic info
                'revision no.': 'Revision No.',
                'revision': 'Revision No.',
                'date': 'Date',
                
                # Vendor info
                'vendor code': 'Vendor Code',
                'code': 'Vendor Code',
                'vendor name': 'Vendor Name',
                'name': 'Vendor Name',
                'vendor location': 'Vendor Location',
                'location': 'Vendor Location',
                
                # Part info
                'part no.': 'Part No.',
                'part number': 'Part No.',
                'part description': 'Part Description',
                'description': 'Part Description',
                'part unit weight': 'Part Unit Weight',
                'unit weight': 'Part Unit Weight',
                'weight': 'Part Unit Weight',
                'part l': 'Part L',
                'length': 'Part L',
                'part w': 'Part W',
                'width': 'Part W',
                'part h': 'Part H',
                'height': 'Part H',

                
                 # INNER packaging - completely separate
                'inner l': 'Inner L',
                'inner l-mm': 'Inner L',
                'inner w': 'Inner W', 
                'inner w-mm': 'Inner W',
                'inner h': 'Inner H',
                'inner h-mm': 'Inner H',
                'inner qty/pack': 'Inner Qty/Pack',
                'inner empty weight': 'Inner Empty Weight',
                'inner pack weight': 'Inner Pack Weight',
                
                # PRIMARY packaging - separate from inner
                'primary packaging type': 'Primary Packaging Type',
                'primary l': 'Primary L',
                'primary l-mm': 'Primary L-mm',
                'primary w': 'Primary W-mm',
                'primary w-mm': 'Primary W',
                'primary h': 'Primary H-mm',
                'primary h-mm': 'Primary H',
                'primary qty/pack': 'Primary Qty/Pack',
                'primary empty weight': 'Primary Empty Weight',
                'primary pack weight': 'Primary Pack Weight',
            
                # Generic packaging (when not specified as primary or inner)
                'packaging type': 'Packaging Type',
                'qty/pack': 'Qty/Pack',
                'empty weight': 'Empty Weight',
                'pack weight': 'Pack Weight',
               
                # Secondary packaging
                'secondary packaging type': 'Secondary Packaging Type',
                'secondary l-mm': 'Secondary L-mm',
                'secondary l': 'Secondary L-mm',
                'secondary w-mm': 'Secondary W-mm',
                'secondary w': 'Secondary W-mm',
                'secondary h-mm': 'Secondary H-mm',
                'secondary h': 'Secondary H-mm',
                'secondary qty/pack': 'Secondary Qty/Pack',
                'secondary empty weight': 'Secondary Empty Weight',
                'secondary pack weight': 'Secondary Pack Weight',
                
                # Additional procedure parameters
                'qty/veh': 'Qty/Veh',
                'qty per vehicle': 'Qty/Veh',
                'layer': 'Layer',
                'layers': 'Layer',
                'level': 'Level',
                'levels': 'Level',
                
                # Approval
                'issued by': 'Issued By',
                'reviewed by': 'Reviewed By',
                'approved by': 'Approved By',
                
                # Additional
                'problem if any': 'Problem If Any',
                'caution': 'Caution'
            }
            
            # Extract data from DataFrame
            for col in df.columns:
                col_lower = str(col).lower().strip()
                if col_lower in field_mapping:
                    field_name = field_mapping[col_lower]
                    # Get first non-null value from the column
                    values = df[col].dropna()
                    if len(values) > 0:
                        extracted_data[field_name] = str(values.iloc[0])
                        
            # Debug: Print what was extracted
            print("=== EXTRACTED DATA ===")
            for key, value in extracted_data.items():
                if any(x in key.lower() for x in ['inner', 'primary', 'secondary', 'qty']):
                    print(f"{key}: {value}")
            print("=====================")
            
            # Try to extract procedure steps if they exist
            for i in range(1, 12):  # Updated to 11 steps
                step_patterns = [f'procedure step {i}', f'step {i}', f'{i}']
                for col in df.columns:
                    col_lower = str(col).lower().strip()
                    if any(pattern in col_lower for pattern in step_patterns):
                        values = df[col].dropna()
                        if len(values) > 0:
                            extracted_data[f'Procedure Step {i}'] = str(values.iloc[0])
                        break
            
            st.success(f"Successfully extracted {len(extracted_data)} fields from Excel file")
            return extracted_data
            
        except Exception as e:
            st.error(f"Error reading Excel file: {str(e)}")
            return {}
    
    def extract_images_from_excel(self, uploaded_file):
        """Extract images from Excel file based on column headers and row positions"""
        images_data = {
            'Current Packaging': None,
            'Primary Packaging': None,
            'Secondary Packaging': None,
            'Label': None
        }
        tmp_file_path = None
        try:
            # Save uploaded file to temporary location
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                tmp_file_path = tmp_file.name
        
            # Load workbook and extract images
            wb = load_workbook(tmp_file_path)
            ws = wb.active

            # Find header positions (search in first few rows)
            header_positions = {}
            header_row = None
        
            # Search for headers in the first 10 rows
            for row_idx in range(1, 11):
                row_headers_found = 0
                temp_positions = {}
            
                for col_idx in range(1, ws.max_column + 1):
                    try:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            cell_value = str(cell_value).strip().lower()
                        
                            # More flexible header matching
                            if any(keyword in cell_value for keyword in ["current packaging", "current pack"]):
                                temp_positions['Current Packaging'] = col_idx - 1  # Convert to 0-based
                                row_headers_found += 1
                            elif any(keyword in cell_value for keyword in ["primary packaging", "primary pack"]):
                                temp_positions['Primary Packaging'] = col_idx - 1
                                row_headers_found += 1
                            elif any(keyword in cell_value for keyword in ["secondary packaging", "secondary pack"]):
                                temp_positions['Secondary Packaging'] = col_idx - 1
                                row_headers_found += 1
                            elif "label" in cell_value:
                                temp_positions['Label'] = col_idx - 1
                                row_headers_found += 1
                    except Exception:
                        continue
                # If we found multiple headers in this row, it's likely the header row
                if row_headers_found >= 2:
                    header_positions = temp_positions
                    header_row = row_idx
                    break
            if not header_positions:
                st.warning("⚠️ Could not find column headers in the Excel file")
                return images_data
            # Process images if they exist
            if hasattr(ws, '_images') and ws._images:
                for idx, img in enumerate(ws._images):
                    try:
                        # Convert image to PIL Image
                        image_stream = io.BytesIO(img._data())
                        pil_image = PILImage.open(image_stream)
                    
                        # Get anchor position
                        anchor = img.anchor
                        col_idx = None
                        row_idx = None
                    
                        # Get position from anchor
                        if hasattr(anchor, '_from') and anchor._from:
                            col_idx = anchor._from.col  # 0-based
                            row_idx = anchor._from.row + 1  # Convert to 1-based for comparison
                        elif hasattr(anchor, 'col') and hasattr(anchor, 'row'):
                            col_idx = anchor.col
                            row_idx = anchor.row + 1
                        if col_idx is not None and row_idx is not None:
                            # Only consider images that are BELOW the header row
                            if header_row and row_idx > header_row:
                                # Find the closest matching header column
                                best_match = None
                                min_distance = float('inf')
                            
                                for category, expected_col in header_positions.items():
                                    distance = abs(col_idx - expected_col)
                                    if distance < min_distance:
                                        min_distance = distance
                                        best_match = category
                                # Assign to best match if within reasonable distance (allow 1-2 column difference)
                                if best_match and min_distance <= 2:
                                    # Special handling: Current and Primary packaging should have same image
                                    if best_match == 'Current Packaging':
                                        images_data['Current Packaging'] = pil_image
                                        # Also assign to Primary Packaging if it doesn't have an image yet
                                        if not images_data['Primary Packaging']:
                                            images_data['Primary Packaging'] = pil_image
                                    elif best_match == 'Primary Packaging':
                                        images_data['Primary Packaging'] = pil_image
                                        # Also assign to Current Packaging if it doesn't have an image yet
                                        if not images_data['Current Packaging']:
                                            images_data['Current Packaging'] = pil_image
                                    else:
                                        # For Secondary Packaging and Label, assign normally
                                        images_data[best_match] = pil_image
                                else:
                                    # Fallback: assign based on column order
                                    sorted_headers = sorted(header_positions.items(), key=lambda x: x[1])
                                    if len(sorted_headers) > 0:
                                        # Find which header this image is closest to
                                        for i, (category, _) in enumerate(sorted_headers):
                                            if not images_data[category]:
                                                images_data[category] = pil_image
                                                break
                    except Exception as img_error:
                        # Silently continue if there's an error with an individual image
                        continue
            # If Current and Primary are still empty but we have images, try a simpler approach
            if not any(images_data.values()) and hasattr(ws, '_images') and ws._images:
                # Simple fallback: assign first few images to categories in order
                categories = ['Current Packaging', 'Primary Packaging', 'Secondary Packaging', 'Label']
                for idx, img in enumerate(ws._images[:len(categories)]):
                    try:
                        image_stream = io.BytesIO(img._data())
                        pil_image = PILImage.open(image_stream)
                    
                        category = categories[idx]
                        images_data[category] = pil_image
                    
                        # If assigning to Current, also assign to Primary (they should be same)
                        if category == 'Current Packaging':
                            images_data['Primary Packaging'] = pil_image
                        elif category == 'Primary Packaging':
                            images_data['Current Packaging'] = pil_image
                    except Exception:
                        continue
            return images_data
        except Exception as e:
            st.error(f"❌ Could not extract images: {str(e)}")
            return images_data
        finally:
            # Clean up temporary file
            if tmp_file_path and os.path.exists(tmp_file_path):
                try:
                    os.unlink(tmp_file_path)
                except Exception:
                    pass

    def apply_border_to_range(self, ws, start_cell, end_cell):
        """Apply borders to a range of cells"""
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Parse cell references
        start_col = ord(start_cell[0]) - ord('A')
        start_row = int(start_cell[1:])
        end_col = ord(end_cell[0]) - ord('A')
        end_row = int(end_cell[1:])
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col+1)
                cell.border = border
    
    def add_image_to_cell_range(self, ws, pil_image, start_cell, end_cell):
        """Add PIL image to specified cell range in worksheet with proper sizing"""
        try:
            # Convert PIL image to bytes
            img_buffer = io.BytesIO()
            pil_image.save(img_buffer, format='PNG')
            img_buffer.seek(0)
        
            # Create openpyxl Image
            img = Image(img_buffer)
        
            # Parse cell coordinates
            start_col_letter = start_cell[0]
            start_row = int(start_cell[1:])
            end_col_letter = end_cell[0]
            end_row = int(end_cell[1:])
        
            # Convert column letters to numbers
            start_col_num = ord(start_col_letter.upper()) - ord('A') + 1
            end_col_num = ord(end_col_letter.upper()) - ord('A') + 1
        
            # Calculate total width and height based on cell dimensions
            total_width = 0
            for col_num in range(start_col_num, end_col_num + 1):
                col_letter = chr(ord('A') + col_num - 1)
                # Get column width (default Excel column width is ~8.43 characters = ~64 pixels)
                col_width = ws.column_dimensions[col_letter].width or 12  # Default to 12 if not set
                # Convert Excel column width to pixels (approximate: 1 character ≈ 7.5 pixels)
                total_width += col_width * 7.5
            total_height = 0
            for row_num in range(start_row, end_row + 1):
                # Get row height (Excel default is ~15 points = ~20 pixels)
                row_height = ws.row_dimensions[row_num].height or 16  # Default to 16 if not set
                # Convert points to pixels (1 point ≈ 1.33 pixels)
                total_height += row_height * 1.33
        
            # Add some padding (reduce by 10% to ensure it fits within borders)
            total_width *= 0.9
            total_height *= 0.9
        
            # Maintain aspect ratio while fitting within the cell range
            original_width, original_height = pil_image.size
            aspect_ratio = original_width / original_height
        
            # Calculate scaling factors
            width_scale = total_width / original_width
            height_scale = total_height / original_height
        
            # Use the smaller scale to maintain aspect ratio and fit within bounds
            scale = min(width_scale, height_scale)
        
            # Apply the scaling
            img.width = int(original_width * scale)
            img.height = int(original_height * scale)
        
            # Add image to worksheet at the start cell
            ws.add_image(img, start_cell)
        
            print(f"Image added to {start_cell}:{end_cell} with dimensions {img.width}x{img.height}")
            return True
        
        except Exception as e:
            print(f"Error adding image to cell range {start_cell}:{end_cell}: {e}")
            return False


    # Alternative method with more precise cell dimension calculation
    def add_image_to_cell_range_precise(self, ws, pil_image, start_cell, end_cell):
        """Add PIL image to specified cell range with more precise dimension calculation"""
        try:
            from openpyxl.utils import column_index_from_string, get_column_letter
        
            # Convert PIL image to bytes
            img_buffer = io.BytesIO()
            pil_image.save(img_buffer, format='PNG')
            img_buffer.seek(0)
        
            # Create openpyxl Image
            img = Image(img_buffer)
        
            # Parse cell coordinates more precisely
            start_col_idx = column_index_from_string(start_cell.split(start_cell.lstrip('ABCDEFGHIJKLMNOPQRSTUVWXYZ'))[0])
            start_row = int(start_cell.lstrip('ABCDEFGHIJKLMNOPQRSTUVWXYZ'))
            end_col_idx = column_index_from_string(end_cell.split(end_cell.lstrip('ABCDEFGHIJKLMNOPQRSTUVWXYZ'))[0])
            end_row = int(end_cell.lstrip('ABCDEFGHIJKLMNOPQRSTUVWXYZ'))
        
            # Calculate total dimensions more precisely
            total_width_chars = 0
            for col_idx in range(start_col_idx, end_col_idx + 1):
                col_letter = get_column_letter(col_idx)
                col_width = ws.column_dimensions[col_letter].width or 8.43  # Excel default
                total_width_chars += col_width
        
            total_height_points = 0
            for row_num in range(start_row, end_row + 1):
                row_height = ws.row_dimensions[row_num].height or 15  # Excel default
                total_height_points += row_height
        
            # Convert to pixels (more accurate conversion)
            # 1 character width ≈ 7 pixels, 1 point ≈ 1.333 pixels
            total_width_pixels = total_width_chars * 7 * 0.9  # 10% padding
            total_height_pixels = total_height_points * 1.333 * 0.9  # 10% padding
        
            # Maintain aspect ratio
            original_width, original_height = pil_image.size
        
            # Calculate scaling to fit within the cell range
            width_ratio = total_width_pixels / original_width
            height_ratio = total_height_pixels / original_height
            scale_ratio = min(width_ratio, height_ratio)
        
            # Apply scaling
            img.width = int(original_width * scale_ratio)
            img.height = int(original_height * scale_ratio)
        
            # Add image to worksheet
            ws.add_image(img, start_cell)
        
            print(f"Precise image sizing: {start_cell}:{end_cell} -> {img.width}x{img.height}px")
            print(f"Cell range: {total_width_chars:.1f} chars x {total_height_points:.1f} pts")
        
            return True
        
        except Exception as e:
            print(f"Error in precise image placement: {e}")
            # Fallback to simpler method
            return self.add_image_to_cell_range(ws, pil_image, start_cell, end_cell)


    # Enhanced method with specific handling for your template's cell ranges
    def add_image_to_template_cell_range(self, ws, pil_image, start_cell, end_cell):
        """Optimized for the specific cell ranges in your packaging template"""
        try:
            # Convert PIL image to bytes
            img_buffer = io.BytesIO()
            pil_image.save(img_buffer, format='PNG')
            img_buffer.seek(0)
        
            # Create openpyxl Image
            img = Image(img_buffer)
        
            # Define specific dimensions for known cell ranges in your template
            cell_range_dimensions = {
                # Primary Packaging area (A32:C37) - 3 cols x 6 rows
                'A37:C42': {'width': 560, 'height': 96},  # 3*12*7*0.9 x 6*16*1.33*0.9
                # Secondary Packaging area (E32:F37) - 2 cols x 6 rows  
                'E37:F42': {'width': 300, 'height': 96},   # 2*12*7*0.9 x 6*16*1.33*0.9
                # Label area (H32:K37) - 4 cols x 6 rows
                'H37:K42': {'width': 560, 'height': 96},   # 4*12*7*0.9 x 6*16*1.33*0.9
                # Current Packaging area (L2:L8) - 1 col x 7 rows (tall)
                'L2:L16': {'width': 675, 'height': 149},    # 30*7*0.9 x 7*16*1.33*0.9
            }
        
            # Create cell range key
            range_key = f"{start_cell}:{end_cell}"
        
            if range_key in cell_range_dimensions:
                # Use predefined dimensions
                target_width = cell_range_dimensions[range_key]['width']
                target_height = cell_range_dimensions[range_key]['height']
            else:
                # Fallback to calculation
                print(f"Unknown cell range {range_key}, calculating dimensions...")
                return self.add_image_to_cell_range(ws, pil_image, start_cell, end_cell)
        
            # Maintain aspect ratio while fitting within target dimensions
            original_width, original_height = pil_image.size
        
            # Calculate scaling factors
            width_scale = target_width / original_width
            height_scale = target_height / original_height
        
            # Use the smaller scale to maintain aspect ratio
            scale = min(width_scale, height_scale)
        
            # Apply the scaling
            img.width = int(original_width * scale)
            img.height = int(original_height * scale)
        
            # Add image to worksheet
            ws.add_image(img, start_cell)
        
            print(f"Template image added: {range_key} -> {img.width}x{img.height}px (scale: {scale:.2f})")
        
            return True
        
        except Exception as e:
            print(f"Error adding template image to {start_cell}:{end_cell}: {e}")
            return False

    def create_exact_template_excel(self):
        """Create the exact Excel template matching the image"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Packaging Instruction"
        
        # Define styles
        blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        light_blue_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=12)
        black_font = Font(color="000000", bold=True, size=14)
        regular_font = Font(color="000000", size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True)
        
        # Set column widths to match the image exactly
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 14
        ws.column_dimensions['K'].width = 14
        ws.column_dimensions['L'].width = 36

        # Set row heights
        for row in range(1, 51):
            ws.row_dimensions[row].height = 16

        # Header Row - "Packaging Instruction"
        ws.merge_cells('A1:K1')
        ws['A1'] = "Packaging Instruction"
        ws['A1'].fill = blue_fill
        ws['A1'].font = white_font
        ws['A1'].alignment = center_alignment
        self.apply_border_to_range(ws, 'A1', 'K1')

        # Current Packaging header (right side)
        ws['L1'] = "CURRENT PACKAGING"
        ws['L1'].fill = blue_fill
        ws['L1'].font = white_font
        ws['L1'].border = border
        ws['L1'].alignment = center_alignment

        # Revision information row
        ws['A2'] = "Revision No."
        ws['A2'].border = border
        ws['A2'].alignment = left_alignment
        ws['A2'].font = header_font

        ws.merge_cells('B2:E2')
        ws['B2'] = "01"
        ws['B2'].border = border
        self.apply_border_to_range(ws, 'B2', 'E2')

        # Date field
        ws['F2'] = "Date"
        ws['F2'].border = border
        ws['F2'].alignment = left_alignment
        ws['F2'].font = header_font

        # Merge cells for date value
        ws.merge_cells('G2:K2')
        ws['G2'] = ""
        ws['G2'].border = border
        self.apply_border_to_range(ws, 'G2', 'K2')

        ws['L2'] = ""
        ws['L2'].border = border

        # Row 4 - Section headers
        ws.merge_cells('A4:D4')
        ws['A4'] = "Vendor Information"
        ws['A4'].font = title_font
        ws['A4'].alignment = center_alignment
        self.apply_border_to_range(ws, 'A4', 'D4')

        ws.merge_cells('F4:I4')
        ws['F4'] = "Part Information"
        ws['F4'].font = title_font
        ws['F4'].alignment = center_alignment
        self.apply_border_to_range(ws, 'F4', 'I4')

        # Apply borders to remaining cells in row 4
        for col in ['J', 'K', 'L']:
            ws[f'{col}4'] = ""
            ws[f'{col}4'].border = border

        # Vendor Code Row
        ws['A5'] = "Code"
        ws['A5'].font = header_font
        ws['A5'].alignment = left_alignment
        ws['A5'].border = border

        ws.merge_cells('B5:D5')
        ws['B5'] = ""
        self.apply_border_to_range(ws, 'B5', 'D5')
        
        # Part fields
        ws['F5'] = "Part No."
        ws['F5'].border = border
        ws['F5'].alignment = left_alignment
        ws['F5'].font = header_font

        ws.merge_cells('G5:K5')
        ws['G5'] = ""
        self.apply_border_to_range(ws, 'G5', 'K5')

        ws['L5'] = ""
        ws['L5'].border = border

        # Vendor Name Row
        ws['A6'] = "Name"
        ws['A6'].font = header_font
        ws['A6'].alignment = left_alignment
        ws['A6'].border = border

        ws.merge_cells('B6:D6')
        ws['B6'] = ""
        self.apply_border_to_range(ws, 'B6', 'D6')

        ws['F6'] = "Description"
        ws['F6'].border = border
        ws['F6'].alignment = left_alignment
        ws['F6'].font = header_font

        ws.merge_cells('G6:K6')
        ws['G6'] = ""
        self.apply_border_to_range(ws, 'G6', 'K6')

        ws['L6'] = ""
        ws['L6'].border = border

        # Vendor Location Row
        ws['A7'] = "Location"
        ws['A7'].font = header_font
        ws['A7'].alignment = left_alignment
        ws['A7'].border = border

        ws.merge_cells('B7:D7')
        ws['B7'] = ""
        self.apply_border_to_range(ws, 'B7', 'D7')

        ws['F7'] = "Unit Weight"
        ws['F7'].border = border
        ws['F7'].alignment = left_alignment
        ws['F7'].font = header_font

        ws.merge_cells('G7:K7')
        ws['G7'] = ""
        self.apply_border_to_range(ws, 'G7', 'K7')

        ws['L7'] = ""
        ws['L7'].border = border

        # Additional row after Unit Weight (Row 8) for L, W, H
        ws['F8'] = "L"
        ws['F8'].border = border
        ws['F8'].alignment = left_alignment
        ws['F8'].font = header_font

        ws['G8'] = ""
        ws['G8'].border = border

        ws['H8'] = "W"
        ws['H8'].border = border
        ws['H8'].alignment = center_alignment
        ws['H8'].font = header_font

        ws['I8'] = ""
        ws['I8'].border = border

        ws['J8'] = "H"
        ws['J8'].border = border
        ws['J8'].alignment = center_alignment
        ws['J8'].font = header_font

        ws['K8'] = ""
        ws['K8'].border = border

        # Empty cells for A-E and L in row 8
        for col in ['A', 'B', 'C', 'D', 'E', 'L']:
            ws[f'{col}8'] = ""
            ws[f'{col}8'].border = border

        # LEAVE ROW 9 EMPTY - BEFORE PRIMARY PACKAGING

        # Title row for Primary Packaging - MOVED TO ROW 10
        ws.merge_cells('A10:K10')
        ws['A10'] = "Primary Packaging Instruction (Primary / Internal)"
        ws['A10'].fill = blue_fill
        ws['A10'].font = white_font
        ws['A10'].alignment = center_alignment
        self.apply_border_to_range(ws, 'A10', 'K10')

        # Primary packaging headers - MOVED TO ROW 11
        headers = ["Packaging Type", "L-mm", "W-mm", "H-mm", "Qty/Pack", "Empty Weight", "Pack Weight"]
        for i, header in enumerate(headers):
            col = chr(ord('A') + i)
            ws[f'{col}11'] = header
            ws[f'{col}11'].border = border
            ws[f'{col}11'].alignment = center_alignment
            ws[f'{col}11'].font = Font(bold=True)  # MADE BOLD

        # Empty cells for remaining columns in row 11
        for col in ['H', 'I', 'J', 'K', 'L']:
            ws[f'{col}11'] = ""
            ws[f'{col}11'].border = border

        # Primary packaging data rows (12-14) - UPDATED ROW NUMBERS
        for row in range(12, 15):
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}{row}'] = ""
                ws[f'{col}{row}'].border = border

        # TOTAL row - UPDATED ROW NUMBER
        ws['D14'] = "TOTAL"
        ws['D14'].border = border
        ws['D14'].font = black_font
        ws['D14'].alignment = center_alignment

        # LEAVE ROW 15 EMPTY - BEFORE SECONDARY PACKAGING

        # Secondary Packaging Instruction header - MOVED TO ROW 16
        ws.merge_cells('A16:K16')
        ws['A16'] = "Secondary Packaging Instruction (Outer / External)"
        ws['A16'].fill = blue_fill
        ws['A16'].font = white_font
        ws['A16'].alignment = center_alignment
        self.apply_border_to_range(ws, 'A16', 'K16')

        ws['L18'] = "PROBLEM IF ANY:"
        ws['L18'].border = border
        ws['L18'].font = black_font
        ws['L18'].alignment = left_alignment

        # Secondary packaging headers - MOVED TO ROW 17
        for i, header in enumerate(headers):
            col = chr(ord('A') + i)
            ws[f'{col}17'] = header
            ws[f'{col}17'].border = border
            ws[f'{col}17'].alignment = center_alignment
            ws[f'{col}17'].font = Font(bold=True)  # MADE BOLD

        # Empty cells for remaining columns in row 17
        for col in ['H', 'I', 'J', 'K']:
            ws[f'{col}17'] = ""
            ws[f'{col}17'].border = border

        ws['L19'] = "CAUTION: REVISED DESIGN"
        ws['L19'].fill = red_fill
        ws['L19'].font = white_font
        ws['L19'].border = border
        ws['L19'].alignment = center_alignment

        # Secondary packaging data rows (18-20) - UPDATED ROW NUMBERS
        for row in range(18, 21):
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}{row}'] = ""
                ws[f'{col}{row}'].border = border

        # TOTAL row for secondary - UPDATED ROW NUMBER
        ws['D20'] = "TOTAL"
        ws['D20'].border = border
        ws['D20'].font = black_font
        ws['D20'].alignment = center_alignment

        # LEAVE ROW 21 EMPTY - BEFORE PACKAGING PROCEDURE

        # Packaging Procedure section - MOVED TO ROW 22
        ws.merge_cells('A22:K22')
        ws['A22'] = "Packaging Procedure"
        ws['A22'].fill = blue_fill
        ws['A22'].font = white_font
        ws['A22'].alignment = center_alignment
        self.apply_border_to_range(ws, 'A22', 'K22')

        # Packaging procedure steps (rows 23-33) - UPDATED ROW NUMBERS
        for i in range(1, 12):
            row = 22 + i
            ws[f'A{row}'] = str(i)
            ws[f'A{row}'].border = border
            ws[f'A{row}'].alignment = center_alignment
            ws[f'A{row}'].font = regular_font

            # MERGE CELLS B to J for each procedure step
            ws.merge_cells(f'B{row}:K{row}')
            ws[f'B{row}'] = ""
            ws[f'B{row}'].alignment = left_alignment
            self.apply_border_to_range(ws, f'B{row}', f'K{row}')

        # LEAVE ROW 34 EMPTY - BEFORE REFERENCE IMAGES

        # Reference Images/Pictures section - MOVED TO ROW 35
        ws.merge_cells('A35:K35')
        ws['A35'] = "Reference Images/Pictures"
        ws['A35'].fill = blue_fill
        ws['A35'].font = white_font
        ws['A35'].alignment = center_alignment
        self.apply_border_to_range(ws, 'A35', 'K35')

        # Image section headers - UPDATED ROW NUMBER
        ws.merge_cells('A36:C36')
        ws['A36'] = "Primary Packaging"
        ws['A36'].alignment = center_alignment
        ws['A36'].font = Font(bold=True)  # MADE BOLD
        self.apply_border_to_range(ws, 'A36', 'C36')

        ws.merge_cells('D36:G36')
        ws['D36'] = "Secondary Packaging"
        ws['D36'].alignment = center_alignment
        ws['D36'].font = Font(bold=True)  # MADE BOLD
        self.apply_border_to_range(ws, 'D36', 'G36')

        ws.merge_cells('H36:K36')
        ws['H36'] = "Label"
        ws['H36'].alignment = center_alignment
        ws['H36'].font = Font(bold=True)  # MADE BOLD
        self.apply_border_to_range(ws, 'H36', 'K36')

        # Image placeholder areas (rows 37-42) - UPDATED ROW NUMBERS
        ws.merge_cells('A37:C42')
        ws['A37'] = "Primary\nPackaging"
        ws['A37'].alignment = center_alignment
        ws['A37'].font = regular_font
        self.apply_border_to_range(ws, 'A37', 'C42')

        # Arrow 1 - UPDATED ROW NUMBER
        ws['D40'] = "→"
        ws['D40'].border = border
        ws['D40'].alignment = center_alignment
        ws['D40'].font = Font(size=20, bold=True)

        # Secondary Packaging image area - UPDATED ROW NUMBERS
        ws.merge_cells('E37:F42')
        ws['E37'] = "SECONDARY\nPACKAGING"
        ws['E37'].alignment = center_alignment
        ws['E37'].font = regular_font
        ws['E37'].fill = light_blue_fill
        self.apply_border_to_range(ws, 'E37', 'F42')

        # Arrow 2 - UPDATED ROW NUMBER
        ws['G40'] = "→"
        ws['G40'].border = border
        ws['G40'].alignment = center_alignment
        ws['G40'].font = Font(size=20, bold=True)

        # Label image area - UPDATED ROW NUMBERS
        ws.merge_cells('H37:K42')
        ws['H37'] = "LABEL"
        ws['H37'].alignment = center_alignment
        ws['H37'].font = regular_font
        self.apply_border_to_range(ws, 'H37', 'K42')

        # Add borders to remaining cells in image section - UPDATED ROW NUMBERS
        for row in range(37, 43):
            for col in ['D', 'G', 'L']:
                if row != 40 or col != 'D':  # Skip D40 which has arrow
                    if row != 40 or col != 'G':  # Skip G40 which has arrow
                        ws[f'{col}{row}'] = ""
                        ws[f'{col}{row}'].border = border

        # LEAVE ROW 43 EMPTY - BEFORE FIRST APPROVAL SECTION

        # First Approval Section - MOVED TO ROW 44
        ws.merge_cells('A44:C44')
        ws['A44'] = "Issued By"
        ws['A44'].alignment = center_alignment
        ws['A44'].font = Font(bold=True)  # MADE BOLD
        self.apply_border_to_range(ws, 'A44', 'C44')

        ws.merge_cells('D44:G44')
        ws['D44'] = "Reviewed By"
        ws['D44'].alignment = center_alignment
        ws['D44'].font = Font(bold=True)  # MADE BOLD
        self.apply_border_to_range(ws, 'D44', 'G44')

        ws.merge_cells('H44:K44')
        ws['H44'] = "Approved By"
        ws['H44'].alignment = center_alignment
        ws['H44'].font = Font(bold=True)  # MADE BOLD
        self.apply_border_to_range(ws, 'H44', 'K44')

        # First signature boxes (rows 45-48) - UPDATED ROW NUMBERS
        ws.merge_cells('A45:C48')
        ws['A45'] = ""
        self.apply_border_to_range(ws, 'A45', 'C48')

        ws.merge_cells('D45:G48')
        ws['D45'] = ""
        self.apply_border_to_range(ws, 'D45', 'G48')

        ws.merge_cells('H45:K48')
        ws['H45'] = ""
        self.apply_border_to_range(ws, 'H45', 'K48')

        # LEAVE ROW 49 EMPTY - BEFORE SECOND APPROVAL SECTION

        # Second Approval Section - MOVED TO ROW 50
        ws.merge_cells('A50:C50')
        ws['A50'] = "Issued By"
        ws['A50'].alignment = center_alignment
        ws['A50'].font = Font(bold=True)  # MADE BOLD
        self.apply_border_to_range(ws, 'A50', 'C50')

        ws.merge_cells('D50:G50')
        ws['D50'] = "Reviewed By"
        ws['D50'].alignment = center_alignment
        ws['D50'].font = Font(bold=True)  # MADE BOLD
        self.apply_border_to_range(ws, 'D50', 'G50')

        ws.merge_cells('H50:K50')
        ws['H50'] = "Approved By"
        ws['H50'].alignment = center_alignment
        ws['H50'].font = Font(bold=True)  # MADE BOLD
        self.apply_border_to_range(ws, 'H50', 'K50')

        # Second signature boxes (rows 51-54) - UPDATED ROW NUMBERS
        ws.merge_cells('A51:C54')
        ws['A51'] = ""
        self.apply_border_to_range(ws, 'A51', 'C54')

        ws.merge_cells('D51:G54')
        ws['D51'] = ""
        self.apply_border_to_range(ws, 'D51', 'G54')

        ws.merge_cells('H51:K54')
        ws['H51'] = ""
        self.apply_border_to_range(ws, 'H51', 'K54')

        # Return the workbook
        return wb
    
    def populate_template_with_data(self, wb, data_dict, procedures_list=None, images_data=None):
        """Populate the template with data from dictionary and optional procedures"""
        ws = wb.active
        # Map data to specific cells
        cell_mapping = {
            'Revision No.': 'B2',
            'Date': 'G2',
            'Vendor Code': 'B5',
            'Vendor Name': 'B6',
            'Vendor Location': 'B7',
            'Part No.': 'G5',
            'Part Description': 'G6',
            'Part Unit Weight': 'G7',
            'Part L': 'G8',
            'Part W': 'I8',
            'Part H': 'K8',
            # Updated Primary packaging fields - NEW ROW NUMBERS
            'Primary Packaging Type': 'A12',  # Was A11
            'Primary L-mm': 'B12',           # Was B11
            'Primary W-mm': 'C12',           # Was C11
            'Primary H-mm': 'D12',           # Was D11
            'Primary Qty/Pack': 'E12',       # Was E11
            'Primary Empty Weight': 'F12',   # Was F11
            'Primary Pack Weight': 'G12',    # Was G11
            # Secondary packaging - NEW ROW NUMBERS
            'Secondary Packaging Type': 'A18', # Was A16
            'Secondary L-mm': 'B18',          # Was B16
            'Secondary W-mm': 'C18',          # Was C16
            'Secondary H-mm': 'D18',          # Was D16
            'Secondary Qty/Pack': 'E18',      # Was E16
            'Secondary Empty Weight': 'F18',  # Was F16
            'Secondary Pack Weight': 'G18',   # Was G16
            'Problem If Any': 'L19',          # Was L17
            'Issued By': 'A45',               # Was A40
            'Reviewed By': 'D45',             # Was D40
            'Approved By': 'H45',             # Was H40
            'Caution': 'L20'                  # Was L18
        }
        # Populate cells with data
        for field, cell in cell_mapping.items():
            if field in data_dict and data_dict[field]:
                try:
                    ws[cell] = data_dict[field]
                except Exception as e:
                    print(f"Error populating cell {cell} with field {field}: {e}")
                    print(f"Field value: {data_dict[field]}")
                    print(f"Field type: {type(data_dict[field])}")
        
        # Handle procedure steps from data_dict (updated to 11 steps) - UPDATED ROW NUMBERS
        try:
            for i in range(1, 12):  # Updated to handle 11 procedures
                procedure_key = f'Procedure Step {i}'
                if procedure_key in data_dict and data_dict[procedure_key]:
                    # Convert to string if it's not already
                    procedure_value = str(data_dict[procedure_key])
                    # Skip if it's a slice object representation
                    if not procedure_value.startswith('slice('):
                        row = 22 + i  # Procedure rows now start from 23 (was 20)
                        ws[f'B{row}'] = procedure_value
                    else:
                        print(f"Skipping {procedure_key} - contains slice object: {procedure_value}")
        except Exception as e:
            print(f"Error handling procedure steps from data_dict: {e}")
        
        # Populate procedures if provided as separate list - UPDATED ROW NUMBERS
        if procedures_list:
            try:
                # Ensure procedures_list is actually a list and not a slice object
                if isinstance(procedures_list, list):
                    for i, procedure in enumerate(procedures_list[:11]):  # Updated to max 11 procedures
                        if procedure and str(procedure).strip():  # Only add non-empty procedures
                            # Convert to string and check it's not a slice
                            procedure_str = str(procedure)
                            if not procedure_str.startswith('slice('):
                                row = 23 + i  # Procedure rows now start from 23 (was 20)
                                ws[f'B{row}'] = procedure_str
                            else:
                                print(f"Skipping procedure {i+1} - contains slice object: {procedure_str}")
                else:
                    print(f"procedures_list is not a list. Type: {type(procedures_list)}, Value: {procedures_list}")
            except Exception as e: 
                print(f"Error handling procedures_list: {e}")
                print(f"procedures_list type: {type(procedures_list)}")
                print(f"procedures_list value: {procedures_list}")
        
        # Handle images if provided - UPDATED CELL REFERENCES
        if images_data:
            try:
                # Add images to specific cell ranges - UPDATED ROW NUMBERS
                if images_data.get('Primary Packaging'):
                    self.add_image_to_cell_range(ws, images_data['Primary Packaging'], 'A37', 'C42')  # Was A32:C37
                if images_data.get('Secondary Packaging'):
                    self.add_image_to_cell_range(ws, images_data['Secondary Packaging'], 'E37', 'F42')  # Was E32:F37
                if images_data.get('Label'):
                    self.add_image_to_cell_range(ws, images_data['Label'], 'H37', 'K42')  # Was H32:K37
                if images_data.get('Current Packaging'):
                    self.add_image_to_cell_range(ws, images_data['Current Packaging'], 'L2', 'L17')
            except Exception as e:
                print(f"Error handling images: {e}")
        
        return wb
        
def main():
    st.set_page_config(page_title="Exact Packaging Template Generator", layout="wide")
    st.title("🏭 Packaging Instruction Template Generator")
    st.markdown("Upload and modify existing packaging instruction templates")
    
    # Initialize template manager
    template_manager = ExactPackagingTemplateManager()
    
    # Main content - Upload & Modify Existing
    st.header("📁 Upload & Modify Existing Template")
    uploaded_file = st.file_uploader(
        "Upload Existing Excel Template",
        type=['xlsx', 'xls'],
        help="Upload an existing packaging template to extract and modify data"
    )
    
    if uploaded_file is not None:
        st.success("File uploaded successfully!")
    
        # Extract data and images from uploaded file
        with st.spinner("Extracting data from Excel file..."):
            extracted_data = template_manager.extract_data_from_excel(uploaded_file)

            # ✅ ADD DEBUG HERE
            print("=== DEBUG PLACEHOLDERS ===")
            for key in ['Qty/Veh', 'Layer', 'Level', 'Inner L', 'Inner W', 'Inner H', 'Inner Qty/Pack']:
                print(f"{key}: {extracted_data.get(key)}")
            print("==========================")
        
            # Reset file pointer for image extraction
            uploaded_file.seek(0)
            extracted_images = template_manager.extract_images_from_excel(uploaded_file)
        
            # Show quick summary of what was extracted
            col1, col2 = st.columns(2)
            with col1:
                extracted_count = sum(1 for v in extracted_data.values() if v)
                st.metric("Data Fields Extracted", extracted_count)
            with col2:
                images_count = sum(1 for v in extracted_images.values() if v)
                st.metric("Images Extracted", images_count)
    
        if extracted_data:
            st.subheader("📊 Extracted Data")
            with st.expander("View Extracted Fields", expanded=False):
                for key, value in extracted_data.items():
                    if value:
                        st.write(f"**{key}**: {value}")
                        
            # Packaging procedures section
            st.subheader("📋 Update Packaging Procedures")
        
            col1, col2 = st.columns([1, 2])
        
            with col1:
                st.write("**Select Packaging Type:**")
                procedure_type = st.selectbox(
                    "Packaging Procedure Type",
                    ["Select Packaging Procedure", "BOX IN BOX SENSITIVE", "BOX IN BOX", "CARTON BOX WITH SEPARATOR FOR ONE PART", "INDIVIDUAL NOT SENSITIVE", "INDIVIDUAL PROTECTION FOR EACH PART", "INDIVIDUAL SENSITIVE", "MANY IN ONE TYPE", "SINGLE BOX"],
                    help="Select a packaging type to auto-populate procedure steps"
                )
            with col2:
                if procedure_type and procedure_type != "Select Packaging Procedure":
                    st.info(f"Selected: {procedure_type}")
                    if procedure_type in template_manager.packaging_procedures:
                        procedures = template_manager.get_procedure_steps(procedure_type, extracted_data)
                        st.write("**Procedure Steps Preview:**")
                        for i, step in enumerate(procedures, 1):
                            if step.strip():
                                st.write(f"{i}. {step}")
            
            st.subheader("📁 Generate Updated Template")
        
            if st.button("🚀 Generate Updated Excel Template", type="primary"):
                # Use original extracted data
                updated_form_data = extracted_data.copy()
            
                # Update only the procedure steps if a type is selected
                if procedure_type and procedure_type != "Select Packaging Procedure" and procedure_type in template_manager.packaging_procedures:
                    procedure_steps = template_manager.get_procedure_steps(procedure_type, extracted_data)
                    for i, step in enumerate(procedure_steps, 1):
                        updated_form_data[f'Procedure Step {i}'] = step
                    # Also update the primary packaging type
                    updated_form_data['Primary Packaging Type'] = procedure_type
                    st.success(f"Updated procedures for {procedure_type}")
                    
                # Generate Excel file
                try:
                    wb = template_manager.create_exact_template_excel()
                    wb = template_manager.populate_template_with_data(wb, updated_form_data, None, extracted_images)
                
                    # Save to buffer
                    buffer = io.BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)
                
                    # Provide download
                    st.success("✅ Updated template generated successfully!")
                    st.download_button(
                        label="⬇️ Download Updated Excel Template",
                        data=buffer.getvalue(),
                        file_name=f"Updated_Packaging_Template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"Error generating updated template: {str(e)}")
        else:
            st.warning("Could not extract data from the uploaded file. Please check the file format and try again.")
    else:
        # Show instructions when no file is uploaded
        st.info("👆 Please upload an Excel template file to get started")
        st.markdown("""
        **Instructions:**
        1. Upload your existing packaging template Excel file
        2. Review the extracted data
        3. Select a packaging procedure type to update the template
        4. Download the updated template
        """)
        
if __name__ == "__main__":
    main()
